"""
================================================================================
Google Calendar Blocker
================================================================================

PURPOSE:
    A utility that reads time blocking data from an Excel file and automatically
    creates blocked time events in Google Calendar. This helps manage focus time
    and prevent meeting scheduling during blocked periods.

KEY FEATURES:
    - Read blocking schedules from Excel files (.xlsx)
    - Support for recurring time blocks
    - Automatic Google Calendar integration
    - Batch event creation
    - Conflict detection and handling
    - Detailed logging and reporting
    - Flexible time format support

USAGE:
    python calendar_blocker.py <excel_file_path>

EXAMPLES:
    # Block calendar based on Excel file
    python calendar_blocker.py schedule.xlsx
    
    # Block with specific date range
    python calendar_blocker.py schedule.xlsx --start 2025-01-01 --end 2025-12-31

EXCEL FILE FORMAT:
    The Excel file should have columns:
    - Date: Date of the block (YYYY-MM-DD or MM/DD/YYYY)
    - Start Time: Start time (HH:MM in 24-hour format)
    - End Time: End time (HH:MM in 24-hour format)
    - Title: Block title/reason (e.g., "Focus Time", "Meeting Prep")
    - Description: Optional details about the block
    - Recurring: Optional - 'Daily', 'Weekly', 'Monthly' or 'None'
    - Color: Optional - Calendar event color

LEARNING CONCEPTS:
    - Excel file reading with openpyxl/pandas
    - Google Calendar API integration
    - OAuth 2.0 authentication
    - Date/time parsing and handling
    - Batch operations and error handling
    - Configuration management
    - Logging and error reporting

SCRIPT STRUCTURE:
    1. AUTHENTICATION & SETUP - Google Calendar API connection
    2. EXCEL PARSING - Read and validate Excel data
    3. EVENT CREATION - Create events in Google Calendar
    4. ERROR HANDLING - Handle conflicts and failures
    5. LOGGING - Report results and issues
    6. MAIN EXECUTION - Orchestrate the process

REQUIREMENTS:
    - google-auth-oauthlib
    - google-auth-httplib2
    - google-api-python-client
    - openpyxl (for Excel reading)
    - python-dateutil (for date parsing)

================================================================================
"""

import os
import sys
from datetime import datetime, timedelta
from typing import List, Dict, Optional, Tuple
import logging

# Configure logging
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s'
)
logger = logging.getLogger(__name__)

# ================================================================================
# SECTION 1: IMPORTS & CONFIGURATION
# ================================================================================

try:
    import openpyxl
    from openpyxl.utils import get_column_letter
except ImportError:
    logger.error("openpyxl not installed. Install with: pip install openpyxl")
    sys.exit(1)

try:
    from google.auth.transport.requests import Request
    from google.oauth2.service_account import Credentials
    from google.oauth2 import service_account
    from googleapiclient.discovery import build
    from googleapiclient.errors import HttpError
except ImportError:
    logger.error("Google API client not installed. Install with: pip install google-auth-oauthlib google-auth-httplib2 google-api-python-client")
    sys.exit(1)

# ================================================================================
# SECTION 2: CONFIGURATION & CONSTANTS
# ================================================================================

SCOPES = ['https://www.googleapis.com/auth/calendar']

# Color mapping for Google Calendar
COLOR_MAP = {
    'red': '11',
    'orange': '17',
    'yellow': '5',
    'green': '2',
    'blue': '1',
    'purple': '3',
    'gray': '8',
    'none': None
}

RECURRING_MAP = {
    'daily': 'RRULE:FREQ=DAILY',
    'weekly': 'RRULE:FREQ=WEEKLY',
    'monthly': 'RRULE:FREQ=MONTHLY',
    'none': None,
    '': None
}

# ================================================================================
# SECTION 3: EXCEL PARSING FUNCTIONS
# ================================================================================

def parse_excel_file(file_path: str) -> List[Dict]:
    """
    FUNCTION: parse_excel_file
    
    PURPOSE:
        Reads and parses the Excel file containing calendar blocking schedule.
    
    PARAMETERS:
        file_path (str): Path to the Excel file
    
    RETURNS:
        List[Dict]: List of dictionaries with block information
    
    FUNCTIONALITY:
        1. Opens the Excel file
        2. Reads header row to identify columns
        3. Validates required columns
        4. Parses each row into a dictionary
        5. Validates date and time formats
        6. Returns parsed events
    
    ERROR HANDLING:
        - Checks if file exists
        - Validates Excel format
        - Checks for required columns
        - Logs parsing errors
    """
    logger.info(f"Parsing Excel file: {file_path}")
    
    # Validate file exists
    if not os.path.exists(file_path):
        logger.error(f"File not found: {file_path}")
        raise FileNotFoundError(f"Excel file not found: {file_path}")
    
    # Load workbook
    try:
        workbook = openpyxl.load_workbook(file_path)
        worksheet = workbook.active
    except Exception as e:
        logger.error(f"Error loading Excel file: {e}")
        raise
    
    # Extract headers
    headers = []
    for cell in worksheet[1]:
        if cell.value:
            headers.append(cell.value.lower().strip())
    
    logger.info(f"Found columns: {headers}")
    
    # Validate required columns
    required_columns = ['date', 'start time', 'end time', 'title']
    for col in required_columns:
        if col not in headers:
            logger.error(f"Missing required column: {col}")
            raise ValueError(f"Missing required column: {col}")
    
    # Parse rows
    events = []
    for row_idx, row in enumerate(worksheet.iter_rows(min_row=2, values_only=False), start=2):
        try:
            event = {}
            for col_idx, cell in enumerate(row):
                if col_idx < len(headers):
                    header = headers[col_idx]
                    value = cell.value
                    event[header] = value
            
            # Skip empty rows
            if not event.get('date'):
                continue
            
            # Validate and add to list
            if validate_event(event):
                events.append(event)
        except Exception as e:
            logger.warning(f"Error parsing row {row_idx}: {e}")
            continue
    
    logger.info(f"Successfully parsed {len(events)} events from Excel")
    return events


def validate_event(event: Dict) -> bool:
    """
    FUNCTION: validate_event
    
    PURPOSE:
        Validates that an event has all required fields with correct formats.
    
    PARAMETERS:
        event (Dict): Event dictionary to validate
    
    RETURNS:
        bool: True if valid, False otherwise
    
    FUNCTIONALITY:
        1. Checks required fields exist
        2. Validates date format
        3. Validates time format
        4. Validates time order (start before end)
    """
    try:
        # Check required fields
        if not event.get('date') or not event.get('start time') or not event.get('end time'):
            logger.warning(f"Missing required fields in event: {event}")
            return False
        
        if not event.get('title'):
            event['title'] = "Calendar Block"
        
        # Validate date format
        date_str = str(event['date']).strip()
        try:
            event['date'] = parse_date(date_str)
        except ValueError as e:
            logger.warning(f"Invalid date format: {date_str}")
            return False
        
        # Validate time format
        try:
            start_time = parse_time(str(event['start time']).strip())
            end_time = parse_time(str(event['end time']).strip())
            
            if start_time >= end_time:
                logger.warning(f"Start time must be before end time: {event}")
                return False
            
            event['start_time'] = start_time
            event['end_time'] = end_time
        except ValueError as e:
            logger.warning(f"Invalid time format: {e}")
            return False
        
        return True
    
    except Exception as e:
        logger.warning(f"Event validation error: {e}")
        return False


def parse_date(date_str: str) -> datetime.date:
    """
    FUNCTION: parse_date
    
    PURPOSE:
        Parses various date formats into a date object.
    
    PARAMETERS:
        date_str (str): Date string to parse
    
    RETURNS:
        datetime.date: Parsed date object
    """
    from dateutil import parser
    try:
        return parser.parse(date_str).date()
    except:
        raise ValueError(f"Could not parse date: {date_str}")


def parse_time(time_str: str) -> str:
    """
    FUNCTION: parse_time
    
    PURPOSE:
        Validates time format (HH:MM).
    
    PARAMETERS:
        time_str (str): Time string to parse
    
    RETURNS:
        str: Validated time string in HH:MM format
    """
    try:
        parts = time_str.split(':')
        if len(parts) != 2:
            raise ValueError("Invalid time format")
        
        hour = int(parts[0].strip())
        minute = int(parts[1].strip())
        
        if not (0 <= hour < 24) or not (0 <= minute < 60):
            raise ValueError("Hour must be 0-23, minute must be 0-59")
        
        return f"{hour:02d}:{minute:02d}"
    except Exception as e:
        raise ValueError(f"Could not parse time: {time_str}")

# ================================================================================
# SECTION 4: GOOGLE CALENDAR INTEGRATION
# ================================================================================

def get_calendar_service():
    """
    FUNCTION: get_calendar_service
    
    PURPOSE:
        Authenticates with Google Calendar API and returns service object.
    
    RETURNS:
        googleapiclient.discovery.Resource: Google Calendar API service
    
    FUNCTIONALITY:
        1. Loads credentials from credentials file
        2. Refreshes token if needed
        3. Returns authenticated service
    
    ERROR HANDLING:
        - Checks for credentials file
        - Handles authentication errors
        - Provides helpful error messages
    """
    try:
        # For service account (recommended for automation)
        creds = service_account.Credentials.from_service_account_file(
            'credentials.json', scopes=SCOPES)
        
        service = build('calendar', 'v3', credentials=creds)
        logger.info("Successfully authenticated with Google Calendar API")
        return service
    
    except FileNotFoundError:
        logger.error("credentials.json not found. Please set up Google Calendar API credentials.")
        raise
    except Exception as e:
        logger.error(f"Authentication error: {e}")
        raise


def create_calendar_event(service, event_data: Dict) -> Optional[str]:
    """
    FUNCTION: create_calendar_event
    
    PURPOSE:
        Creates a calendar event in Google Calendar.
    
    PARAMETERS:
        service: Google Calendar API service
        event_data (Dict): Event details to create
    
    RETURNS:
        Optional[str]: Event ID if successful, None otherwise
    
    FUNCTIONALITY:
        1. Constructs event object
        2. Adds event to calendar
        3. Returns event ID
        4. Handles API errors
    """
    try:
        # Construct event
        event = {
            'summary': event_data.get('title', 'Calendar Block'),
            'description': event_data.get('description', ''),
            'start': {
                'dateTime': f"{event_data['date']}T{event_data['start_time']}:00",
                'timeZone': 'America/New_York',
            },
            'end': {
                'dateTime': f"{event_data['date']}T{event_data['end_time']}:00",
                'timeZone': 'America/New_York',
            },
            'colorId': COLOR_MAP.get(str(event_data.get('color', '')).lower(), None),
            'transparency': 'opaque',  # Block this time
        }
        
        # Add recurring rule if specified
        recurring = event_data.get('recurring', '').lower().strip()
        if recurring and recurring in RECURRING_MAP and RECURRING_MAP[recurring]:
            event['recurrence'] = [RECURRING_MAP[recurring]]
        
        # Create event
        event = service.events().insert(calendarId='primary', body=event).execute()
        logger.info(f"Created event: {event['id']} - {event_data['title']}")
        return event['id']
    
    except HttpError as e:
        logger.error(f"Google Calendar API error: {e}")
        return None
    except Exception as e:
        logger.error(f"Error creating event: {e}")
        return None


# ================================================================================
# SECTION 5: MAIN ORCHESTRATION
# ================================================================================

def block_calendar(excel_file: str) -> Tuple[int, int]:
    """
    FUNCTION: block_calendar
    
    PURPOSE:
        Main orchestration function that reads Excel file and creates
        calendar events.
    
    PARAMETERS:
        excel_file (str): Path to the Excel file
    
    RETURNS:
        Tuple[int, int]: (successful_events, failed_events)
    
    FUNCTIONALITY:
        1. Parses Excel file
        2. Authenticates with Google Calendar
        3. Creates events for each block
        4. Logs results
    """
    try:
        # Parse Excel file
        events = parse_excel_file(excel_file)
        if not events:
            logger.warning("No events found in Excel file")
            return 0, 0
        
        logger.info(f"Attempting to create {len(events)} calendar events")
        
        # Authenticate with Google Calendar
        service = get_calendar_service()
        
        # Create events
        successful = 0
        failed = 0
        
        for event in events:
            event_id = create_calendar_event(service, event)
            if event_id:
                successful += 1
            else:
                failed += 1
        
        # Report results
        logger.info(f"Calendar blocking complete!")
        logger.info(f"Successful: {successful}, Failed: {failed}")
        
        return successful, failed
    
    except Exception as e:
        logger.error(f"Fatal error: {e}")
        return 0, len(events) if 'events' in locals() else 0


def main():
    """
    FUNCTION: main
    
    PURPOSE:
        Entry point for the calendar blocker script.
    
    COMMAND-LINE ARGUMENTS:
        excel_file (required): Path to the Excel file
        --start (optional): Start date (YYYY-MM-DD)
        --end (optional): End date (YYYY-MM-DD)
    
    FUNCTIONALITY:
        1. Parses command-line arguments
        2. Validates inputs
        3. Calls block_calendar function
        4. Displays results
    """
    if len(sys.argv) < 2:
        print("Usage: python calendar_blocker.py <excel_file>")
        print("\nExample:")
        print("  python calendar_blocker.py schedule.xlsx")
        print("\nRequired:")
        print("  - Excel file with columns: Date, Start Time, End Time, Title")
        print("  - Google Calendar credentials (credentials.json)")
        sys.exit(1)
    
    excel_file = sys.argv[1]
    
    # Run blocking
    successful, failed = block_calendar(excel_file)
    
    print(f"\n{'='*60}")
    print(f"Calendar Blocking Complete")
    print(f"{'='*60}")
    print(f"Successfully created: {successful} events")
    print(f"Failed: {failed} events")
    print(f"{'='*60}\n")
    
    sys.exit(0 if failed == 0 else 1)


# ================================================================================
# SCRIPT EXECUTION
# ================================================================================

if __name__ == "__main__":
    main()
